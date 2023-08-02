from selenium import webdriver
import openpyxl
import time

# открываем файл Excel
workbook = openpyxl.load_workbook('RDB.xlsx')
ws = workbook.active

# задания из файла в список
tasks = []
for row in ws.iter_rows(min_row=2, values_only=True):
    tasks.append((row[0], row[1]))


# Запускаем хром
driver = webdriver.Chrome('chromedriver.exe')

# Перебираем задания
for i, task in enumerate(tasks):
    task_id, status = task

    # Заходим на страницу нашего id задания
    driver.get(f'http://t2ru-rollapp-t1.corp.tele2.ru/p/form.aspx?op=form&k=c3a5t34r{task_id}')


    try:
        # Кликаем на кнопку "Вернуться"
        return_button = driver.find_element_by_class_name('glyph glyphBack')
        # return_button = driver.find_element_by_css_selector("[title='Вернуться']")
        return_button.click()

        # Если нет ошибок, записываем "успешно" в файл
        ws.cell(row=i + 2, column=2).value = "успешно"

    except:
        # Если  ошибка, записываем "ошибка" в файл
        ws.cell(row=i + 2, column=2).value = "ошибка"

# Сейвим файл Excel
workbook.save('RDB.xlsx')


driver.quit()